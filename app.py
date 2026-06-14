"""
IBKR Portfolio Dashboard - Flask App v2
Strategy-based portfolio management with collapsible strategy groups.
"""

from flask import Flask, render_template, request, jsonify, session, redirect
from werkzeug.utils import secure_filename
import pandas as pd
from pathlib import Path
import json
import os
import subprocess
import re
from datetime import datetime

from parser import (
    detect_format, parse_positions, parse_trades,
    parse_transaction_history, apply_transactions_to_portfolio,
)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key-change-in-prod")
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "ibkr2026")
app.config["UPLOAD_FOLDER"] = Path(__file__).parent / "uploads"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

DATA_DIR = Path(__file__).parent / "data"
PORTFOLIO_FILE = DATA_DIR / "portfolio.json"
TARGET_FILE = DATA_DIR / "target_allocation.json"

# === IBKR 佣金费率配置（阶梯式 Tiered）===
COMMISSION_CONFIG = {
    # 阶梯式佣金 (Tiered)
    "stock_per_share": 0.0035,      # 美股每股 $0.0035（<10万股/月）
    "stock_min_per_order": 0.35,    # 最低 $0.35/笔
    "stock_max_pct": 0.005,        # 最高交易额的 0.5%
    # 期权阶梯式
    "option_per_contract": 0.65,   # 每份合约 $0.65（<10万合约/月）
    "option_min_per_order": 0.35,   # 最低 $0.35/笔
    # SEC/FINRA/TAF 等监管费（正股卖出时）
    "sec_fee_rate": 0.0000278,     # SEC fee ~0.00278% of sell amount
    "taf_per_share": 0.000166,     # TAF (FINRA) ~$0.000166/share
    "finra_taf_min": 0.01,         # FINRA TAF 最低 $0.01
}

def calc_commission(quantity, price, is_option=False, contracts=None):
    """计算单笔交易佣金（含税费）
    
    参数:
        quantity: 股数（正股）
        price: 每股/每股期权价格
        is_option: 是否期权
        contracts: 合约数（期权时使用）
    返回: {commission: float, fees: float, total_cost: float}
    """
    if is_option and contracts:
        # 期权佣金
        commission = max(COMMISSION_CONFIG["option_min_per_order"],
                        COMMISSION_CONFIG["option_per_contract"] * contracts)
        return {"commission": round(commission, 2), "fees": 0, "total_cost": round(commission, 2)}
    
    # 正股佣金
    trade_value = abs(quantity) * price
    commission_by_share = COMMISSION_CONFIG["stock_per_share"] * abs(quantity)
    commission_by_pct = trade_value * COMMISSION_CONFIG["stock_max_pct"]
    commission = max(COMMISSION_CONFIG["stock_min_per_order"],
                     min(commission_by_share, commission_by_pct))
    
    # 监管费（仅在卖出时收取，这里统一按卖出计算，买入时caller传0）
    # 不在这里自动加监管费，由调用方决定是否卖出
    
    return {"commission": round(commission, 2), "fees": 0, "total_cost": round(commission, 2)}

def calc_sell_fees(quantity, price):
    """计算正股卖出的SEC/TAF监管费"""
    trade_value = abs(quantity) * price
    sec_fee = max(0.01, trade_value * COMMISSION_CONFIG["sec_fee_rate"])
    taf_fee = max(COMMISSION_CONFIG["finra_taf_min"], 
                 abs(quantity) * COMMISSION_CONFIG["taf_per_share"])
    return round(sec_fee + taf_fee, 2)


DEFAULT_STRATEGIES = [
    {"id": "dca", "name": "定投仓 (DCA)", "icon": "📊", "color": "#6366f1",
     "desc": "定期定额买入个股和大盘ETF", "type": "stock"},
    {"id": "wheel", "name": "轮子策略仓 (Wheel)", "icon": "🎡", "color": "#22c55e",
     "desc": "Sell Put → 被行权 → Covered Call → 卖出", "type": "option"},
    {"id": "leaps", "name": "LEAPS Call仓", "icon": "🚀", "color": "#a855f7",
     "desc": "长期期权（到期1年+）看多策略", "type": "option"},
    {"id": "swing", "name": "波段仓 (Swing)", "icon": "⚡", "color": "#f59e0b",
     "desc": "短线波段交易", "type": "stock"},
    {"id": "cash", "name": "现金仓", "icon": "💵", "color": "#3b82f6",
     "desc": "现金储备，等待机会", "type": "stock"},
]

# ---- Data Layer ----

DEFAULT_PORTFOLIO = {
    "strategies": DEFAULT_STRATEGIES,
    "positions": [
        # DCA positions (普通策略保持不变)
        {"id":"p1","symbol":"AAPL","strategy":"dca","quantity":150,"avg_price":175.5,"current_price":195.2,"market_value":29280,"pnl":2955,"pnl_pct":11.22,"notes":""},
        {"id":"p2","symbol":"MSFT","strategy":"dca","quantity":80,"avg_price":410,"current_price":445.6,"market_value":35648,"pnl":2848,"pnl_pct":8.68,"notes":""},
        {"id":"p3","symbol":"QQQ","strategy":"dca","quantity":50,"avg_price":460,"current_price":498.3,"market_value":24915,"pnl":1915,"pnl_pct":8.33,"notes":"大盘ETF"},
        {"id":"p4","symbol":"VOO","strategy":"dca","quantity":30,"avg_price":520,"current_price":555.8,"market_value":16674,"pnl":1074,"pnl_pct":6.88,"notes":"S&P500 ETF"},
        # Wheel positions (轮子策略 - 期权格式，增加delta字段)
        {"id":"w1","symbol":"AMZN","strategy":"wheel","wheel_type":"sell_put","strike":185,"expiry":"2026-07-18","premium":3.20,"contracts":2,"quantity":200,"stock_price":195.50,"cost_basis":181.80,"current_option_price":0.45,"market_value":39100,"pnl":640,"pnl_pct":1.83,"status":"等待行权","delta":-0.25,"notes":""},
        {"id":"w2","symbol":"GOOGL","strategy":"wheel","wheel_type":"covered_call","strike":185,"expiry":"2026-08-15","premium":4.50,"contracts":4,"quantity":400,"stock_price":178.60,"cost_basis":150.70,"current_option_price":6.20,"market_value":71440,"pnl":1800,"pnl_pct":2.59,"status":"卖Covered Call","delta":-0.35,"notes":""},
        {"id":"w3","symbol":"TSLA","strategy":"wheel","wheel_type":"sell_put","strike":240,"expiry":"2026-06-20","premium":5.80,"contracts":1,"quantity":100,"stock_price":268.40,"cost_basis":234.20,"current_option_price":0.12,"market_value":26840,"pnl":580,"pnl_pct":2.21,"status":"等待行权","delta":-0.15,"notes":""},
        # LEAPS Call positions (LEAPS Call - 期权格式，增加delta字段)
        {"id":"l1","symbol":"NVDA","strategy":"leaps","strike":900,"expiry":"2028-01-20","contracts":2,"quantity":200,"buy_price":45.80,"current_option_price":82.50,"stock_price":1050.30,"market_value":16500,"pnl":7340,"pnl_pct":80.13,"delta":0.65,"notes":"2028 LEAPS Call"},
        {"id":"l2","symbol":"META","strategy":"leaps","strike":600,"expiry":"2027-01-15","contracts":3,"quantity":300,"buy_price":32.50,"current_option_price":58.20,"stock_price":565.80,"market_value":17460,"pnl":3810,"pnl_pct":39.08,"delta":0.55,"notes":"2027 LEAPS Call"},
        # Swing positions (普通策略保持不变)
        {"id":"p9","symbol":"TSLA","strategy":"swing","quantity":15,"avg_price":245,"current_price":268.4,"market_value":4026,"pnl":351,"pnl_pct":9.55,"notes":"短线持有"},
        {"id":"p10","symbol":"AMD","strategy":"swing","quantity":30,"avg_price":155,"current_price":168.2,"market_value":5046,"pnl":396,"pnl_pct":8.52,"notes":""},
    ],
    "cash": 35000,
    "history": [],
    "cash_flows": [],
    "cash_base_usd": 35000,
}

def load_portfolio():
    if PORTFOLIO_FILE.exists():
        data = json.loads(PORTFOLIO_FILE.read_text(encoding="utf-8"))
        migrate_realized_profits(data)
        # 迁移：为旧持仓补 open_date，为 buy_trades 补 date
        for p in data.get("positions", []):
            p.setdefault("open_date", "")
            for t in p.get("buy_trades", []):
                t.setdefault("date", "")
        return data
    # First run: seed with default data
    default = dict(DEFAULT_PORTFOLIO)
    save_portfolio(default)
    return default

def _git_sync():
    """Auto git commit + push data changes (background, non-blocking)."""
    import threading
    def _sync():
        try:
            repo = Path(__file__).parent
            subprocess.run(['git', '-C', str(repo), 'add', 'data/'],
                           capture_output=True, timeout=10)
            result = subprocess.run(
                ['git', '-C', str(repo), 'commit', '-m', f'data: auto-save {datetime.now().strftime("%Y-%m-%d %H:%M")}'],
                capture_output=True, timeout=10)
            if result.returncode != 0:
                return
            token = os.environ.get('GITHUB_TOKEN', '')
            if token:
                subprocess.run(
                    ['git', '-C', str(repo), 'push',
                     f'https://x-access-token:{token}@github.com/ericzhouaussie/ibkr-portfolio-dashboard.git', 'main'],
                    capture_output=True, timeout=30)
            else:
                subprocess.run(['git', '-C', str(repo), 'push', 'origin', 'main'],
                               capture_output=True, timeout=30)
        except Exception:
            pass
    threading.Thread(target=_sync, daemon=True).start()


def save_portfolio(data):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PORTFOLIO_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    _git_sync()


def add_cash_flow(amount, currency, rate, note="", flow_type="deposit"):
    """
    flow_type: 'deposit' (入金) or 'withdraw' (出金)
    amount: 金额（正数）
    currency: 'CNY' or 'USD'
    rate: USD/CNY 汇率（currency='CNY' 时需要）
    Returns: (success, new_cash_base_usd)
    """
    portfolio = load_portfolio()
    if currency == 'CNY':
        usd = round(amount / rate, 2) if rate else 0
    else:
        usd = round(amount, 2)
    flow = {
        "id": generate_id("cf_"),
        "type": flow_type,
        "currency": currency,
        "original_amount": amount,
        "rate": rate if currency == 'CNY' else None,
        "amount_usd": usd,
        "note": note,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    portfolio.setdefault("cash_flows", []).append(flow)
    if flow_type == "deposit":
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) + usd, 2)
    else:
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) - usd, 2)
    save_portfolio(portfolio)
    return True, portfolio["cash_base_usd"]


def migrate_cash_flows(portfolio):
    """迁移旧数据：没有 currency 字段的旧记录默认为 CNY"""
    flows = portfolio.get("cash_flows", [])
    changed = False
    for f in flows:
        if "currency" not in f:
            f["currency"] = "CNY"
            f["original_amount"] = f.get("amount_cny", 0)
            changed = True
    if changed:
        save_portfolio(portfolio)


def migrate_realized_profits(portfolio):
    """从历史记录中补充计算已实现盈利（首次初始化）"""
    if "realized_profits" in portfolio:
        return  # 已有数据，跳过
    rp = {}
    for rec in portfolio.get("history", []):
        strat = rec.get("strategy", "")
        pnl = rec.get("pnl", 0)
        if strat:
            rp[strat] = round(rp.get(strat, 0) + pnl, 2)
    if rp:
        portfolio["realized_profits"] = {k: round(v, 2) for k, v in rp.items()}
        save_portfolio(portfolio)


def generate_id(prefix=""):
    """Generate unique ID with optional prefix"""
    import uuid
    return prefix + str(uuid.uuid4())[:8]


def add_realized_profit(portfolio, strategy, pnl):
    """记录已实现盈利到策略累计账户"""
    rp = portfolio.setdefault("realized_profits", {})
    rp[strategy] = round(rp.get(strategy, 0) + pnl, 2)


def get_realized_profits(portfolio):
    """返回每个策略的已实现盈利字典 {strategy_id: amount}"""
    rp = portfolio.get("realized_profits", {})
    # 从历史记录中补充计算（兼容旧数据）
    history = portfolio.get("history", [])
    result = dict(rp)
    for rec in history:
        strat = rec.get("strategy", "")
        pnl = rec.get("pnl", 0)
        # 检查是否已通过 add_realized_profit 记录过
        if strat not in result:
            result[strat] = 0.0
        # 避免 double-counting（如果 realized_profits 中已有，跳过）
        # 旧记录没有这个逻辑，直接累加
    return {k: round(v, 2) for k, v in result.items()}

def load_target_allocation():
    if TARGET_FILE.exists():
        return json.loads(TARGET_FILE.read_text(encoding="utf-8"))
    return None

def save_target_allocation(targets):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TARGET_FILE.write_text(json.dumps(targets, indent=2, ensure_ascii=False), encoding="utf-8")

# ---- Price Fetcher (Twelve Data) ----

def fetch_price_sina(symbol):
    """Fetch latest US stock price via Sina Finance API (no key needed)."""
    try:
        sym_lower = symbol.lower()
        cmd = ['curl', '-s', '--max-time', '8',
               f'https://hq.sinajs.cn/list=gb_{sym_lower}',
               '-H', 'Referer: https://finance.sina.com.cn']
        result = subprocess.run(cmd, capture_output=True)
        if not result.stdout:
            return None
        text = result.stdout.decode('latin-1')
        if '="' in text:
            content = text.split('="')[1].rstrip('";\n')
            fields = content.split(',')
            if len(fields) > 1:
                price = float(fields[1])
                if price > 0:
                    return price
    except:
        pass
    return None


def fetch_price_batch_sina(symbols):
    """Batch fetch US stock prices via Sina Finance API (single request)."""
    try:
        sym_list = ','.join(f'gb_{s.lower()}' for s in symbols)
        cmd = ['curl', '-s', '--max-time', '10',
               f'https://hq.sinajs.cn/list={sym_list}',
               '-H', 'Referer: https://finance.sina.com.cn']
        result = subprocess.run(cmd, capture_output=True)
        if not result.stdout:
            return {}
        text = result.stdout.decode('latin-1')
        prices = {}
        for line in text.strip().split('\n'):
            if '="' not in line:
                continue
            var_name = line.split('=')[0]
            sym = var_name.replace('hq_str_gb_', '').upper()
            content = line.split('="')[1].rstrip('";')
            fields = content.split(',')
            if len(fields) > 1:
                try:
                    price = float(fields[1])
                    if price > 0:
                        prices[sym] = price
                except:
                    pass
        return prices
    except:
        return {}


def refresh_all_prices(api_key=""):
    """Refresh prices for all positions via Sina Finance (no API key needed)."""
    portfolio = load_portfolio()
    positions = portfolio.get("positions", [])
    if not positions:
        return portfolio

    # Deduplicate symbols
    symbols = []
    seen = set()
    for p in positions:
        sym = p.get("symbol", "").upper()
        if sym and sym not in seen:
            symbols.append(sym)
            seen.add(sym)

    # Batch fetch all symbols in one request
    prices = fetch_price_batch_sina(symbols)

    updated = 0
    errors = []
    for p in positions:
        sym = p.get("symbol", "").upper()
        price = prices.get(sym)
        if price is None:
            # Fallback: single fetch
            price = fetch_price_sina(sym)
        if price is None:
            errors.append(sym)
            continue

        strat_type = get_strategy_type(portfolio.get("strategies", []), p.get("strategy", ""))
        if strat_type == "stock":
            p["current_price"] = price
            qty = p.get("quantity", 0)
            avg = p.get("avg_price", 0)
            p["market_value"] = round(abs(qty) * price, 2)
            p["pnl"] = round((price - avg) * qty, 2)
            p["pnl_pct"] = round((price / avg - 1) * 100, 2) if avg else 0
        else:
            # 期权策略：更新正股价
            p["stock_price"] = price
            # 不自动计算盈亏（依赖用户手动更新期权价），仅更新股价
            updated += 1

    save_portfolio(portfolio)
    return {"portfolio": portfolio, "updated": updated, "errors": errors}


# ---- Auth ----

@app.before_request
def require_auth():
    if request.path == "/login" or request.path.startswith("/static/"):
        return
    if not session.get("authenticated"):
        return redirect("/login")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == DASHBOARD_PASSWORD:
            session["authenticated"] = True
            return redirect("/")
        return render_template("login.html", error="密码错误")
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ---- Routes ----

@app.route("/")
def index():
    portfolio = load_portfolio()
    targets = load_target_allocation()
    return render_template("dashboard.html", portfolio=portfolio, targets=targets)


@app.route("/api/portfolio", methods=["GET"])
def get_portfolio():
    return jsonify(load_portfolio())


@app.route("/api/portfolio/order", methods=["POST"])
def save_strategy_order():
    """保存策略排序顺序"""
    data = request.get_json()
    order = data.get("strategies", [])  # 按新顺序排列的策略ID列表
    portfolio = load_portfolio()
    strategies = portfolio.get("strategies", [])
    if not order or not strategies:
        return jsonify({"error": "无效数据"}), 400
    # 按新顺序重排
    id_to_strat = {s["id"]: s for s in strategies}
    portfolio["strategies"] = [id_to_strat[sid] for sid in order if sid in id_to_strat]
    save_portfolio(portfolio)
    return jsonify({"success": True})


def get_strategy_type(strategies, strat_id):
    """获取策略类型，向后兼容硬编码策略ID"""
    strat = next((s for s in strategies if s["id"] == strat_id), None)
    if strat and strat.get("type"):
        return strat["type"]
    # 向后兼容：默认策略ID映射
    if strat_id in ("wheel", "leaps"):
        return "option"
    return "stock"


@app.route("/api/portfolio/position", methods=["POST"])
def add_position():
    data = request.get_json()
    portfolio = load_portfolio()

    position = {
        "id": data.get("id", ""),
        "symbol": data.get("symbol", "").upper(),
        "strategy": data.get("strategy", "dca"),
        "notes": data.get("notes", ""),
        "open_date": data.get("open_date", ""),
    }

    # 根据策略类型处理不同字段
    is_option = get_strategy_type(portfolio.get("strategies", []), position["strategy"]) == "option"
    if is_option:
        # 期权策略字段（通用）
        position["strike"] = float(data.get("strike", 0))
        position["expiry"] = data.get("expiry", "")
        position["option_type"] = data.get("option_type", "put")  # put 或 call
        position["contracts"] = int(data.get("contracts", 1))
        position["quantity"] = position["contracts"] * 100
        position["stock_price"] = float(data.get("stock_price", 0))
        # premium = 用户输入的单张权利金（总权利金 = premium * 100 * |contracts|）
        premium_per_share = float(data.get("premium", 0))
        position["premium"] = premium_per_share
        position["buy_price"] = premium_per_share
        position["current_option_price"] = premium_per_share
        position["delta"] = float(data.get("delta", 0))
        # 判断方向：正数=买期权，负数=卖期权
        is_sold = position["contracts"] < 0
        abs_contracts = abs(position["contracts"])
        # 市场价值 = 当前期权价 * 100 * 合约数（卖期权时为负值表示负债）
        position["market_value"] = round(premium_per_share * 100 * abs_contracts, 2)
        # 盈亏初始化为0（开仓时）
        position["pnl"] = 0
        position["pnl_pct"] = 0
        # 向后兼容 wheel_type
        if is_sold:
            if position["option_type"] == "put":
                position["wheel_type"] = "sell_put"
            else:
                position["wheel_type"] = "sell_call"
            position["status"] = "等待行权"
        # LEAPS 等买期权场景
        if not is_sold and position.get("strategy") == "leaps":
            position["wheel_type"] = "leaps_long"
    else:
        # 正股策略字段 (DCA, Swing)
        position["quantity"] = float(data.get("quantity", 0))
        position["avg_price"] = float(data.get("avg_price", 0))
        position["current_price"] = float(data.get("current_price", data.get("avg_price", 0)))
        qty = position["quantity"]
        ap = position["avg_price"]
        cp = position["current_price"]
        position["market_value"] = round(abs(qty) * cp, 2)
        position["pnl"] = round((cp - ap) * qty, 2)
        position["pnl_pct"] = round((cp / ap - 1) * 100, 2) if ap else 0

    # 新建持仓：补 open_date
    if not position.get("open_date"):
        position["open_date"] = __import__("datetime").datetime.now().strftime("%Y-%m-%d")

    # Update or add (DCA/Swing同策略同symbol自动合并)
    if position["id"]:
        for i, p in enumerate(portfolio["positions"]):
            if p["id"] == position["id"]:
                # 保留 open_date（表单未传时继承旧值）
                if not position.get("open_date") and p.get("open_date"):
                    position["open_date"] = p["open_date"]
                portfolio["positions"][i] = position
                break
        else:
            portfolio["positions"].append(position)
    elif not is_option:
        # 查找同策略下同symbol的持仓，合并（仅正股策略）
        merged = False
        for i, p in enumerate(portfolio["positions"]):
            if (p["strategy"] == position["strategy"]
                    and p.get("symbol", "").upper() == position["symbol"].upper()
                    and "avg_price" in p):
                # 加权平均成本
                old_qty = p["quantity"]
                old_avg = p["avg_price"]
                new_qty = position["quantity"]
                new_avg = position["avg_price"]
                total_qty = old_qty + new_qty
                p["quantity"] = total_qty
                p["avg_price"] = round((old_qty * old_avg + new_qty * new_avg) / total_qty, 2)
                p["current_price"] = position["current_price"]
                p["market_value"] = round(abs(total_qty) * p["current_price"], 2)
                p["pnl"] = round((p["current_price"] - p["avg_price"]) * total_qty, 2)
                p["pnl_pct"] = round((p["current_price"] / p["avg_price"] - 1) * 100, 2) if p["avg_price"] else 0
                # Swing: append to buy_trades for FIFO
                if position["strategy"] == "swing":
                    p.setdefault("buy_trades", []).append({
                        "date": __import__("datetime").datetime.now().strftime("%Y-%m-%d"),
                        "qty": new_qty,
                        "price": new_avg,
                    })
                # 记录加仓到history
                history = portfolio.setdefault("history", [])
                history.append({
                    "id": "hist_" + str(__import__("uuid").uuid4().hex[:8]),
                    "date": __import__("datetime").datetime.now().strftime("%Y-%m-%d"),
                    "symbol": position["symbol"],
                    "action": "BUY",
                    "quantity": new_qty,
                    "price": new_avg,
                    "cost_price": 0,
                    "pnl": 0,
                    "strategy": position["strategy"],
                    "note": f"{position['strategy'].upper()}加仓 {new_qty}股 @{new_avg:.2f}",
                    "commission": calc_commission(new_qty, new_avg)["total_cost"],
                    "fees": 0,
                })
                merged = True
                position = p
                break
        if not merged:
            import uuid
            position["id"] = str(uuid.uuid4())[:8]
            # Swing: initialize buy_trades
            if position["strategy"] == "swing":
                position["buy_trades"] = [{
                    "date": __import__("datetime").datetime.now().strftime("%Y-%m-%d"),
                    "qty": position["quantity"],
                    "price": position["avg_price"],
                }]
            portfolio["positions"].append(position)
    else:
        import uuid
        position["id"] = str(uuid.uuid4())[:8]
        portfolio["positions"].append(position)

    # === 资金来源：现金还是已实现盈利 ===
    profit_source = data.get("profit_source", "")  # 如 "wheel" 或 "swing"
    rp = portfolio.get("realized_profits", {})

    # 开仓现金管理：根据合约数正负号判断买卖方向
    if is_option:
        premium_per = position.get("premium", 0)
        contracts_val = position.get("contracts", 0)
        abs_c = abs(contracts_val)
        is_sold_opt = contracts_val < 0
        if premium_per > 0 and abs_c > 0:
            if is_sold_opt:
                # 卖期权 = 收入权利金
                cash_change = premium_per * abs_c * 100
                portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) + cash_change, 2)
            else:
                # 买期权 = 支付权利金 + 佣金
                cost = premium_per * abs_c * 100
                comm = calc_commission(0, premium_per, is_option=True, contracts=abs_c)
                net_cost = cost + comm["total_cost"]
                if profit_source and rp.get(profit_source, 0) >= net_cost:
                    # 用已实现盈利支付
                    rp[profit_source] = round(rp[profit_source] - net_cost, 2)
                    portfolio["realized_profits"] = rp
                    position["funded_by_profit"] = profit_source
                else:
                    portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) - net_cost, 2)
    else:
        # 正股策略：买入股票/ETF付钱
        qty = position.get("quantity", 0)
        price = position.get("avg_price", 0) or position.get("current_price", 0)
        if qty > 0 and price > 0:
            cost = qty * price
            comm = calc_commission(qty, price)
            net_cost = cost + comm["total_cost"]
            if profit_source and rp.get(profit_source, 0) >= net_cost:
                # 用已实现盈利支付
                rp[profit_source] = round(rp[profit_source] - net_cost, 2)
                portfolio["realized_profits"] = rp
                position["funded_by_profit"] = profit_source
            else:
                portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) - net_cost, 2)

    save_portfolio(portfolio)
    return jsonify({"success": True, "position": position, "realized_profits": portfolio.get("realized_profits", {}), "portfolio": portfolio})


@app.route("/api/portfolio/position/<pos_id>", methods=["DELETE"])
def delete_position(pos_id):
    portfolio = load_portfolio()
    portfolio["positions"] = [p for p in portfolio["positions"] if p["id"] != pos_id]
    save_portfolio(portfolio)
    return jsonify({"success": True, "portfolio": portfolio})


@app.route("/api/portfolio/cash/adjust", methods=["POST"])
def adjust_cash():
    """手动调整现金余额（直接设置新值，不影响出入金流水）"""
    data = request.get_json()
    new_cash = float(data.get("cash_usd", 0))
    portfolio = load_portfolio()
    portfolio["cash_base_usd"] = round(new_cash, 2)
    save_portfolio(portfolio)
    return jsonify({"success": True, "cash_base_usd": portfolio["cash_base_usd"]})


# ---- Cash Flow ----

@app.route("/api/realized-profits", methods=["GET"])
def get_realized_profits():
    """获取各策略已实现盈利"""
    portfolio = load_portfolio()
    return jsonify({"realized_profits": portfolio.get("realized_profits", {}), "portfolio": portfolio})


@app.route("/api/realized-profits", methods=["POST"])
def reset_realized_profit():
    """重置某个策略的已实现盈利（可选）"""
    data = request.get_json()
    portfolio = load_portfolio()
    strategy = data.get("strategy", "")
    if strategy:
        rp = portfolio.get("realized_profits", {})
        if strategy in rp:
            rp[strategy] = 0.0
            portfolio["realized_profits"] = rp
            save_portfolio(portfolio)
    return jsonify({"success": True, "realized_profits": portfolio.get("realized_profits", {})})


@app.route("/api/cash-flow", methods=["GET", "POST"])
def cash_flow():
    """GET: 获取流水列表 / POST: 新增入金/出金记录"""
    portfolio = load_portfolio()
    migrate_cash_flows(portfolio)
    portfolio.setdefault('cash_flows', [])
    if request.method == 'POST':
        data = request.get_json()
        currency = data.get('currency', 'CNY')
        amount = float(data.get('amount', 0))
        rate = float(data.get('rate', 7.25)) if currency == 'CNY' else None
        note = data.get('note', '')
        flow_type = data.get('type', 'deposit')
        if amount <= 0:
            return jsonify({"error": "金额必须大于0"}), 400
        success, new_cash = add_cash_flow(amount, currency, rate, note, flow_type)
        portfolio = load_portfolio()
        return jsonify({"success": True, "cash_base_usd": portfolio['cash_base_usd'], "flows": portfolio['cash_flows']})
    return jsonify({"flows": portfolio['cash_flows'], "cash_base_usd": portfolio.get('cash_base_usd', 0)})


@app.route("/api/cash-flow/<flow_id>", methods=["DELETE"])
def delete_cash_flow(flow_id):
    portfolio = load_portfolio()
    flows = portfolio.get('cash_flows', [])
    flow = next((f for f in flows if f['id'] == flow_id), None)
    if flow:
        if flow['type'] == 'deposit':
            portfolio['cash_base_usd'] = round(portfolio.get('cash_base_usd', 0) - flow['amount_usd'], 2)
        else:
            portfolio['cash_base_usd'] = round(portfolio.get('cash_base_usd', 0) + flow['amount_usd'], 2)
        portfolio['cash_flows'] = [f for f in flows if f['id'] != flow_id]
        save_portfolio(portfolio)
        return jsonify({"success": True, "cash_base_usd": portfolio['cash_base_usd']})
    return jsonify({"error": "记录不存在"}), 404


@app.route("/api/portfolio/transactions/upload", methods=["POST"])
def upload_transactions():
    """Upload IBKR transaction history CSV (中文交易历史)."""
    portfolio = load_portfolio()
    if "file" not in request.files:
        return jsonify({"error": "请选择文件"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "文件名为空"}), 400

    filename = secure_filename(file.filename)
    filepath = app.config["UPLOAD_FOLDER"] / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    file.save(filepath)

    try:
        transactions = parse_transaction_history(str(filepath))
        portfolio, new_tx = apply_transactions_to_portfolio(transactions, portfolio)
        save_portfolio(portfolio)
        return jsonify({
            "success": True,
            "transactions_count": len(new_tx),
            "total_transactions": len(portfolio.get("transactions", [])),
            "cash_base_usd": portfolio.get("cash_base_usd", 0),
            "message": f"成功导入 {len(new_tx)} 条新交易记录"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/portfolio/strategy", methods=["POST"])
def add_strategy():
    data = request.get_json()
    portfolio = load_portfolio()
    strategy = {
        "id": data.get("id", str(__import__("uuid").uuid4())[:8]),
        "name": data.get("name", "新策略"),
        "icon": data.get("icon", "📁"),
        "color": data.get("color", "#8b8d9a"),
        "desc": data.get("desc", ""),
        "type": data.get("type", "stock"),
    }
    portfolio["strategies"].append(strategy)
    save_portfolio(portfolio)
    return jsonify({"success": True, "portfolio": portfolio})

@app.route("/api/portfolio/strategy/<strat_id>", methods=["DELETE"])
def delete_strategy(strat_id):
    portfolio = load_portfolio()
    # 删除策略
    portfolio["strategies"] = [s for s in portfolio["strategies"] if s["id"] != strat_id]
    # 删除该策略下的所有持仓
    portfolio["positions"] = [p for p in portfolio["positions"] if p.get("strategy") != strat_id]
    save_portfolio(portfolio)
    return jsonify({"success": True, "portfolio": portfolio})


@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400

    filename = secure_filename(file.filename)
    filepath = app.config["UPLOAD_FOLDER"] / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    file.save(str(filepath))

    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(filepath, skiprows=1, low_memory=False)
        elif filename.endswith(".xlsx"):
            df = pd.read_excel(filepath, skiprows=1)
        else:
            return jsonify({"error": "仅支持CSV和Excel文件"}), 400

        fmt = detect_format(df)
        upload_type = request.form.get("type", "positions")

        if upload_type == "trades":
            trades = parse_trades(df)
            return jsonify({"success": True, "trades": trades, "count": len(trades)})
        else:
            positions = parse_positions(df, fmt)
            if not positions:
                return jsonify({"error": "未解析到持仓数据，请检查文件格式"}), 400
            # Import into portfolio under specified strategy or DCA default
            strategy = request.form.get("strategy", "dca")
            portfolio = load_portfolio()
            import uuid
            for p in positions:
                p["id"] = str(uuid.uuid4())[:8]
                p["strategy"] = strategy
                p["notes"] = ""
                portfolio["positions"].append(p)
            save_portfolio(portfolio)
            return jsonify({"success": True, "count": len(positions), "portfolio": portfolio})
    except Exception as e:
        return jsonify({"error": "解析失败: %s" % str(e)}), 500


@app.route("/api/refresh-prices", methods=["POST"])
def refresh_prices():
    data = request.get_json() or {}
    api_key = data.get("api_key", "demo")
    result = refresh_all_prices(api_key)
    if isinstance(result, dict) and "portfolio" in result:
        return jsonify({"success": True, "updated": result["updated"], "errors": result["errors"], "portfolio": result["portfolio"]})
    else:
        return jsonify({"success": True, "portfolio": result})


@app.route("/api/option/price", methods=["POST"])
def update_option_price():
    """手动输入期权当前价，计算期权盈亏"""
    data = request.get_json()
    pos_id = data.get("id", "")
    opt_price = data.get("current_option_price", 0)
    if not pos_id:
        return jsonify({"error": "缺少id"}), 400
    try:
        opt_price = float(opt_price) if opt_price else None
    except (ValueError, TypeError):
        opt_price = None

    portfolio = load_portfolio()
    pos = None
    for p in portfolio.get("positions", []):
        if p.get("id") == pos_id:
            pos = p
            break
    if not pos:
        return jsonify({"error": "持仓不存在"}), 404
    if get_strategy_type(portfolio.get("strategies", []), pos["strategy"]) != "option":
        return jsonify({"error": "仅支持期权持仓"}), 400

    # 更新期权当前价
    if opt_price is not None and opt_price >= 0:
        pos["current_option_price"] = opt_price
    else:
        opt_price = pos.get("current_option_price", 0)

    contracts = pos.get("contracts", 1)
    premium = pos.get("premium", pos.get("buy_price", 0))  # 单张权利金
    abs_contracts = abs(contracts)
    is_sold = contracts < 0  # 负数=卖期权
    
    # 卖期权（空头）：盈亏 = (收入权利金 - 当前期权价) * 100 * 合约数
    #   例子：卖Put，收到premium $3，当前价$2 → 盈利$1/张
    #   例子：卖Put，收到premium $3，当前价$5 → 亏损$2/张
    # 买期权（多头）：盈亏 = (当前期权价 - 买入价) * 100 * 合约数
    #   例子：买Call，成本$5，当前价$7 → 盈利$2/张
    if is_sold:
        pos["pnl"] = round((premium - opt_price) * 100 * abs_contracts, 2)
        pos["market_value"] = round(opt_price * 100 * abs_contracts, 2)
        cost = premium * 100 * abs_contracts
        pos["pnl_pct"] = round((pos["pnl"] / cost) * 100, 2) if cost else 0
    else:
        buy_price = premium
        if buy_price > 0:
            pos["pnl"] = round((opt_price - buy_price) * 100 * abs_contracts, 2)
            pos["market_value"] = round(opt_price * 100 * abs_contracts, 2)
            pos["pnl_pct"] = round(((opt_price / buy_price) - 1) * 100, 2) if buy_price else 0

    save_portfolio(portfolio)
    return jsonify({"success": True, "position": pos, "portfolio": portfolio})


@app.route("/api/targets", methods=["GET"])
def get_targets():
    return jsonify({"targets": load_target_allocation() or []})

@app.route("/api/targets", methods=["POST"])
def set_targets():
    data = request.get_json()
    save_target_allocation(data.get("targets", []))
    return jsonify({"success": True})


@app.route("/api/portfolio/position/<pos_id>/close", methods=["POST"])
def close_position(pos_id):
    """平仓API - 移除持仓并创建历史记录"""
    data = request.get_json()  # {"close_price": 0.15}
    portfolio = load_portfolio()
    
    # 查找持仓
    position = None
    position_index = -1
    for i, p in enumerate(portfolio["positions"]):
        if p["id"] == pos_id:
            position = p
            position_index = i
            break
    
    if not position:
        return jsonify({"error": "持仓不存在"}), 404
    
    close_price = float(data.get("close_price", 0))
    
    # 创建历史记录
    history = portfolio.get("history", [])
    
    from datetime import datetime
    close_date = datetime.now().strftime("%Y-%m-%d")
    
    is_option = get_strategy_type(portfolio.get("strategies", []), position["strategy"]) == "option"
    if not is_option:
        return jsonify({"error": "该策略不支持平仓操作"}), 400

    is_wheel = position["strategy"] == "wheel"
    contracts = position["contracts"]
    comm = calc_commission(0, close_price, is_option=True, contracts=contracts)
    
    if is_wheel and position.get("wheel_type"):
        # Wheel策略平仓（特有字段：premium）
        open_premium = position.get("premium", 0)
        pnl = round((open_premium - close_price) * contracts * 100, 2)
        pnl_pct = round((open_premium / close_price - 1) * 100, 2) if close_price > 0 else 0
        
        record = {
            "id": generate_id("h_"),
            "symbol": position["symbol"],
            "strategy": position["strategy"],
            "wheel_type": position.get("wheel_type", ""),
            "strike": position["strike"],
            "expiry": position["expiry"],
            "contracts": contracts,
            "open_premium": open_premium,
            "close_price": close_price,
            "open_delta": position.get("delta", 0),
            "open_date": position.get("open_date", ""),
            "close_date": close_date,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
            "commission": comm["total_cost"],
            "fees": comm["fees"],
            "total_cost": comm["total_cost"],
            "status": "已平仓",
            "notes": ""
        }
        # Wheel平仓：买回期权=付钱
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) + (open_premium - close_price * contracts * 100 - comm["total_cost"]), 2)
    else:
        # 通用期权平仓（LEAPS/自定义期权仓）
        open_price = position.get("buy_price", position.get("premium", 0))
        pnl = round((close_price - open_price) * contracts * 100, 2)
        pnl_pct = round((close_price / open_price - 1) * 100, 2) if open_price > 0 else 0
        
        record = {
            "id": generate_id("h_"),
            "symbol": position["symbol"],
            "strategy": position["strategy"],
            "strike": position["strike"],
            "expiry": position["expiry"],
            "contracts": contracts,
            "open_price": open_price,
            "close_price": close_price,
            "open_delta": position.get("delta", 0),
            "open_date": position.get("open_date", ""),
            "close_date": close_date,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
            "commission": comm["total_cost"],
            "fees": comm["fees"],
            "total_cost": comm["total_cost"],
            "status": "已平仓",
            "notes": ""
        }
        # 平仓卖出期权：收钱 - 佣金
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) + close_price * contracts * 100 - comm["total_cost"], 2)
    
    # 记录已实现盈利
    add_realized_profit(portfolio, position["strategy"], pnl)
    
    # 从历史中移除持仓
    portfolio["positions"].pop(position_index)
    
    # 添加历史记录
    history.append(record)
    portfolio["history"] = history
    
    save_portfolio(portfolio)
    return jsonify({"success": True, "history": record, "realized_profits": portfolio.get("realized_profits", {}), "portfolio": portfolio})


@app.route("/api/history", methods=["GET"])
def get_history():
    """获取所有历史交易记录"""
    portfolio = load_portfolio()
    return jsonify({"history": portfolio.get("history", [])})


@app.route("/api/history", methods=["DELETE"])
def clear_history():
    """清空历史交易记录"""
    portfolio = load_portfolio()
    portfolio["history"] = []
    save_portfolio(portfolio)
    return jsonify({"success": True})


@app.route("/api/export/history")
def export_history():
    """导出交易记录为 Excel 文件"""
    portfolio = load_portfolio()
    history = portfolio.get("history", [])

    if not history:
        return "没有交易记录可导出", 400

    rows = []
    for h in history:
        action = h.get("action", "")
        is_option = h.get("strike") is not None
        is_close = action in ("SELL",) or "close" in str(h.get("status", "")).lower()

        row = {
            "日期": h.get("date", ""),
            "标的": h.get("symbol", ""),
            "操作": "平仓" if is_close else action,
            "策略": h.get("strategy", ""),
            "佣金": h.get("commission", 0),
            "备注": h.get("note", ""),
        }

        # 开仓信息
        if not is_close:
            row["开仓日期"] = h.get("date", "")
            row["开仓价格"] = h.get("price", h.get("premium", h.get("open_premium", "")))
            row["开仓数量"] = h.get("quantity", h.get("contracts", ""))
            row["平仓日期"] = ""
            row["平仓价格"] = ""
            row["平仓数量"] = ""
        else:
            row["开仓日期"] = h.get("open_date", "")
            row["开仓价格"] = h.get("open_price", h.get("open_premium", h.get("cost_price", "")))
            row["开仓数量"] = ""
            row["平仓日期"] = h.get("close_date", h.get("date", ""))
            row["平仓价格"] = h.get("close_price", h.get("price", ""))
            row["平仓数量"] = h.get("quantity", h.get("contracts", ""))

        # 盈亏
        row["盈亏金额"] = h.get("pnl", 0)
        row["盈亏%"] = h.get("pnl_pct", "")

        if is_option:
            row["类型"] = h.get("wheel_type", h.get("option_type", ""))
            row["行权价"] = h.get("strike", "")
            row["到期日"] = h.get("expiry", "")
            row["合约数"] = h.get("contracts", "")
            row["Delta"] = h.get("delta", "")
        else:
            row["成本价"] = h.get("cost_price", "")
            row["费用"] = h.get("fees", 0)

        rows.append(row)

    df = pd.DataFrame(rows)

    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="交易记录")
        ws = writer.sheets["交易记录"]

        # 样式：标题行
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 列宽自适应
        for col_idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
            ws.column_dimensions[col_letter].width = min(max_len, 30)

        # 数据行样式 + 正负色
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_name = df.columns[cell.column - 1]
                if col_name == "盈亏金额" and isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.font = Font(color="16a34a", bold=True)
                    elif cell.value < 0:
                        cell.font = Font(color="dc2626", bold=True)

        ws.freeze_panes = "A2"

    output.seek(0)
    from flask import send_file
    filename = f"IBKR交易记录_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True, download_name=filename)


@app.route("/api/history/<hist_id>", methods=["DELETE"])
def delete_history_item(hist_id):
    """删除单条交易记录，并级联回滚相关影响（现金、已实现盈利、持仓）"""
    portfolio = load_portfolio()
    history = portfolio.get("history", [])
    item = next((h for h in history if h.get("id") == hist_id), None)
    if not item:
        return jsonify({"success": False, "error": "记录不存在"}), 404

    history = [h for h in history if h.get("id") != hist_id]
    portfolio["history"] = history

    action = item.get("action", "")
    strategy = item.get("strategy", "")
    symbol = item.get("symbol", "")
    qty = float(item.get("quantity", 0) or 0)
    price = float(item.get("price", 0) or 0)
    pnl = float(item.get("pnl", 0) or 0)
    comm = float(item.get("commission", 0) or 0)
    fees = float(item.get("fees", 0) or 0)
    total_cost = comm + fees


    if action == "BUY":
        # 回滚：加回现金（持仓成本 + 佣金），删除对应持仓（按symbol/strategy匹配）
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) + qty * price + total_cost, 2)
        if strategy == "dca" or strategy == "swing":
            portfolio["positions"] = [
                p for p in portfolio.get("positions", [])
                if not (p.get("symbol") == symbol and p.get("strategy") == strategy
                       and abs(float(p.get("quantity", 0)) - qty) < 0.01
                       and abs(float(p.get("avg_price", 0)) - price) < 0.01)
            ]

    elif action == "SELL":
        # 回滚：扣除卖出的现金，恢复持仓数量，加回已实现盈利
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) - qty * price + total_cost, 2)
        # 回滚 realized_profits
        rp = portfolio.get("realized_profits", {})
        if strategy in rp and rp[strategy] != 0:
            rp[strategy] = round(rp[strategy] - pnl, 2)
            if abs(rp[strategy]) < 0.01:
                rp[strategy] = 0.0
            portfolio["realized_profits"] = rp
        # 恢复持仓数量
        for p in portfolio.get("positions", []):
            if p.get("symbol") == symbol and p.get("strategy") == strategy:
                p["quantity"] = round(float(p.get("quantity", 0)) + qty, 4)
                if p["quantity"] > 0:
                    # 还原 avg_price（理想情况，实际以close为主）
                    pass
                break

    elif item.get("close_price") is not None:
        # 期权平仓回滚：加回现金（平仓价×合约×100 - 佣金），回滚 realized_profits，恢复持仓
        close_price = float(item.get("close_price", 0))
        contracts = float(item.get("contracts", 0) or 0)
        cp = abs(contracts)
        portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) - close_price * cp * 100 + total_cost, 2)
        rp = portfolio.get("realized_profits", {})
        if strategy in rp and rp[strategy] != 0:
            rp[strategy] = round(rp[strategy] - pnl, 2)
            if abs(rp[strategy]) < 0.01:
                rp[strategy] = 0.0
            portfolio["realized_profits"] = rp
        # 恢复期权持仓
        pos_to_restore = {
            "id": "hist_" + hist_id[:8],
            "symbol": symbol,
            "strategy": strategy,
            "stock_price": float(item.get("stock_price", 0) or 0),
            "strike": float(item.get("strike", 0) or 0),
            "expiry": item.get("expiry", ""),
            "option_type": item.get("option_type", "put"),
            "delta": item.get("open_delta") or item.get("delta"),
            "contracts": float(item.get("contracts", 0) or 0),
            "buy_price": float(item.get("open_price", 0) or item.get("open_premium", 0)),
            "premium": float(item.get("open_premium", 0) or item.get("open_price", 0)),
            "wheel_type": item.get("wheel_type", ""),
            "pnl": 0, "pnl_pct": 0, "current_option_price": float(item.get("open_price", 0) or item.get("open_premium", 0)),
        }
        portfolio.setdefault("positions", []).append(pos_to_restore)


    save_portfolio(portfolio)
    return jsonify({"success": True, "realized_profits": portfolio.get("realized_profits", {}), "cash_base_usd": portfolio.get("cash_base_usd", 0)})


@app.route("/api/portfolio/position/<pos_id>/sell", methods=["POST"])
def sell_position(pos_id):
    """卖出持仓。DCA: 平均成本法; Swing: FIFO逐笔配对。"""
    portfolio = load_portfolio()
    data = request.get_json()
    sell_qty = float(data.get("quantity", 0))
    sell_price = float(data.get("price", 0))

    if sell_qty <= 0 or sell_price <= 0:
        return jsonify({"error": "无效数量或价格"}), 400

    positions = portfolio.get("positions", [])
    pos = next((p for p in positions if p.get("id") == pos_id), None)
    if not pos:
        return jsonify({"error": "持仓不存在"}), 404

    if sell_qty > pos.get("quantity", 0):
        return jsonify({"error": f"卖出数量超过持有量（{pos['quantity']}股）"}), 400

    history = portfolio.get("history", [])
    now_str = __import__("datetime").datetime.now().strftime("%Y-%m-%d")
    symbol = pos.get("symbol", "")
    strategy = pos.get("strategy", "")
    total_pnl = 0

    if strategy == "swing":
        # === FIFO: 逐笔配对 ===
        remaining_sell = sell_qty
        for trade in pos.get("buy_trades", []):
            if remaining_sell <= 0:
                break
            if trade.get("qty", 0) <= 0:
                continue
            match_qty = min(remaining_sell, trade["qty"])
            trade_pnl = (sell_price - trade["price"]) * match_qty
            total_pnl += trade_pnl
            remaining_sell -= match_qty
            trade["qty"] -= match_qty
            # Record per-trade history
            history.append({
                "id": "hist_" + __import__("uuid").uuid4().hex[:8],
                "date": now_str,
                "symbol": symbol,
                "action": "SELL",
                "quantity": match_qty,
                "price": sell_price,
                "cost_price": trade["price"],
                "open_date": trade.get("date", ""),
                "pnl": round(trade_pnl, 2),
                "note": f"Swing卖出 {match_qty}股（FIFO成本${trade['price']:.2f}）",
                "strategy": "swing",
                "commission": calc_commission(match_qty, sell_price)["total_cost"],
                "fees": calc_sell_fees(match_qty, sell_price),
            })
        # Clean empty buy_trades
        pos["buy_trades"] = [t for t in pos.get("buy_trades", []) if t.get("qty", 0) > 0]
    else:
        # === DCA: 平均成本法 ===
        avg_cost = pos.get("avg_price", 0)
        total_pnl = (sell_price - avg_cost) * sell_qty
        history.append({
            "id": "hist_" + __import__("uuid").uuid4().hex[:8],
            "date": now_str,
            "symbol": symbol,
            "action": "SELL",
            "quantity": sell_qty,
            "price": sell_price,
            "cost_price": avg_cost,
            "open_date": pos.get("open_date", ""),
            "pnl": round(total_pnl, 2),
            "note": f"DCA卖出 {sell_qty}股（成本{avg_cost:.2f}）",
            "strategy": "dca",
            "commission": calc_commission(sell_qty, sell_price)["total_cost"],
            "fees": calc_sell_fees(sell_qty, sell_price),
        })

    # Update position
    remaining = pos["quantity"] - sell_qty
    if remaining > 0:
        pos["quantity"] = remaining
        cp = pos.get("current_price", pos.get("avg_price", 0))
        if strategy == "swing" and pos.get("buy_trades"):
            # Recalc avg from remaining buy_trades
            total_cost = sum(t["qty"] * t["price"] for t in pos["buy_trades"])
            pos["avg_price"] = round(total_cost / remaining, 4) if remaining > 0 else 0
        pos["market_value"] = round(remaining * cp, 2)
        pos["pnl"] = round((cp - pos["avg_price"]) * remaining, 2)
        pos["pnl_pct"] = round(((cp - pos["avg_price"]) / pos["avg_price"]) * 100, 2) if pos["avg_price"] > 0 else 0
    else:
        positions = [p for p in positions if p.get("id") != pos_id]
        portfolio["positions"] = positions

    # Cash += sell proceeds - 佣金
    comm = calc_commission(sell_qty, sell_price)
    net_proceeds = sell_qty * sell_price - comm["total_cost"]
    portfolio["cash_base_usd"] = round(portfolio.get("cash_base_usd", 0) + net_proceeds, 2)
    # 记录已实现盈利（每次卖出都计入）
    add_realized_profit(portfolio, strategy, total_pnl)
    portfolio["history"] = history
    save_portfolio(portfolio)
    return jsonify({"success": True, "pnl": round(total_pnl, 2), "remaining": remaining, "realized_profits": portfolio.get("realized_profits", {}), "portfolio": portfolio})


@app.route("/api/backup")
def backup_data():
    """打包下载所有本地数据文件，防止数据丢失"""
    import zipfile, io, os
    from flask import send_file

    portfolio = load_portfolio()
    target = load_target_allocation()

    # 包含所有子目录
    files = {
        "portfolio.json": json.dumps(portfolio, indent=2, ensure_ascii=False),
        "target_allocation.json": json.dumps(target, indent=2, ensure_ascii=False),
    }

    memory_lines = []
    memory_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "memory")
    if os.path.isdir(memory_dir):
        for f in os.listdir(memory_dir):
            if f.endswith(".md"):
                fp = os.path.join(memory_dir, f)
                memory_lines.append((f, open(fp, encoding="utf-8").read()))

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, content in files.items():
            zf.writestr(fname, content)
        if memory_lines:
            zf.writestr("memory/README.txt",
                "此文件夹保存的是 memory/*.md 备份，请放回 memory/ 目录恢复。")
            for fname, content in memory_lines:
                zf.writestr(f"memory/{fname}", content)

    zip_buffer.seek(0)
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ibkr_backup_{ts}.zip"
    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=filename
    )


@app.route("/api/restore", methods=["POST"])
def restore_data():
    """从备份 ZIP 恢复数据"""
    import zipfile, io
    
    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400
    
    file = request.files["file"]
    if not file.filename.endswith(".zip"):
        return jsonify({"error": "请上传 .zip 备份文件"}), 400
    
    try:
        # 读取文件内容到 BytesIO（解决 SpooledTemporaryFile 兼容性问题）
        file_content = file.read()
        zip_buffer = io.BytesIO(file_content)
        zf = zipfile.ZipFile(zip_buffer)
        names = zf.namelist()
        
        # 恢复 portfolio.json
        if "portfolio.json" in names:
            data = json.loads(zf.read("portfolio.json"))
            save_portfolio(data)
        else:
            return jsonify({"error": "备份文件中缺少 portfolio.json"}), 400
        
        # 恢复 target_allocation.json
        if "target_allocation.json" in names:
            target = json.loads(zf.read("target_allocation.json"))
            save_target_allocation(target)
        
        # 恢复 memory 文件
        memory_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "memory")
        for name in names:
            if name.startswith("memory/") and not name.endswith("/"):
                fname = os.path.basename(name)
                if fname == "README.txt":
                    continue
                fp = os.path.join(memory_dir, fname)
                os.makedirs(memory_dir, exist_ok=True)
                with open(fp, "w", encoding="utf-8") as f:
                    f.write(zf.read(name).decode("utf-8"))
        
        portfolio = load_portfolio()
        return jsonify({"success": True, "portfolio": portfolio})
    
    except zipfile.BadZipFile:
        return jsonify({"error": "无效的 ZIP 文件"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def save_target_allocation(data):
    """保存目标配置"""
    with open(TARGET_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)





if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5050))
    app.run(debug=True, port=port, host="0.0.0.0")
