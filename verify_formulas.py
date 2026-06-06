#!/usr/bin/env python3
"""
Verify formulas in the IBKR Portfolio Dashboard.
Reads the Excel file and displays formulas for verification.
"""

from openpyxl import load_workbook

def verify_positions(ws):
    """Verify Positions sheet formulas."""
    print("=" * 60)
    print("Sheet: 持仓录入 (Positions)")
    print("=" * 60)

    print("\nHeaders:")
    for cell in ws[1]:
        if cell.value:
            print(f"  {cell.column_letter}{cell.row}: {cell.value}")

    print("\nSample data formulas (first 3 rows):")
    for row in range(2, 5):
        print(f"\nRow {row}:")
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                print(f"  {cell.column_letter}{row}: {cell.value} (format: {cell.number_format})")

    print("\nEmpty row formulas (row 12):")
    for col in [6, 7, 8]:  # Market Value, P&L Amount, P&L %
        cell = ws.cell(row=12, column=col)
        print(f"  {cell.column_letter}{12}: {cell.value}")


def verify_strategy_summary(ws):
    """Verify Strategy Summary sheet formulas."""
    print("\n" + "=" * 60)
    print("Sheet: 策略汇总 (Strategy Summary)")
    print("=" * 60)

    print("\nCash input cell:")
    print(f"  B3: {ws['B3'].value} (format: {ws['B3'].number_format})")

    print("\nStrategy rows (first 2 strategies):")
    for row in [6, 7]:
        print(f"\nRow {row}:")
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                print(f"  {cell.column_letter}{row}: {cell.value}")

    print("\nTotal row:")
    for col in [1, 3, 4, 5, 6, 7]:
        cell = ws.cell(row=12, column=col)
        if cell.value:
            print(f"  {cell.column_letter}{12}: {cell.value}")

    print("\nCash row:")
    for col in [1, 3, 6]:
        cell = ws.cell(row=13, column=col)
        if cell.value:
            print(f"  {cell.column_letter}{13}: {cell.value}")


def verify_dashboard(ws):
    """Verify Dashboard sheet formulas."""
    print("\n" + "=" * 60)
    print("Sheet: 仪表盘 (Dashboard)")
    print("=" * 60)

    print("\nStatistic cards:")
    for row in [3, 4]:
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                print(f"  {cell.column_letter}{row}: {cell.value}")

    print("\nMillion Challenge Progress:")
    print(f"  B8: {ws['B8'].value} (format: {ws['B8'].number_format})")

    print("\nTarget Allocation Table:")
    for row in range(12, 18):
        print(f"\nRow {row}:")
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                print(f"  {cell.column_letter}{row}: {cell.value}")


def verify_settings(ws):
    """Verify Settings sheet."""
    print("\n" + "=" * 60)
    print("Sheet: 设置 (Settings)")
    print("=" * 60)

    print("\nEditable cells:")
    print(f"  B3 (Cash): {ws['B3'].value} (format: {ws['B3'].number_format})")
    print(f"  B4 (Target): {ws['B4'].value} (format: {ws['B4'].number_format})")


def main():
    """Main verification function."""
    filename = "/Users/ericsteg/.qclaw/workspace/ibkr-portfolio-dashboard/portfolio_dashboard.xlsx"

    print(f"Loading workbook: {filename}")
    wb = load_workbook(filename, data_only=False)  # data_only=False to see formulas

    print("\nSheets:", wb.sheetnames)

    # Verify each sheet
    verify_positions(wb["持仓录入"])
    verify_strategy_summary(wb["策略汇总"])
    verify_dashboard(wb["仪表盘"])
    verify_settings(wb["设置"])

    print("\n" + "=" * 60)
    print("Verification complete!")
    print("=" * 60)
    print("\nNote: To fully verify formula calculations, open the file in Excel")
    print("or install LibreOffice and run recalc.py")


if __name__ == "__main__":
    main()
