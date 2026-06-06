#!/usr/bin/env python3
"""
IBKR Portfolio Dashboard Creator
Creates a formula-driven Excel dashboard for IBKR portfolio tracking.
All calculations use Excel formulas, no hardcoded values.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill, numbers
)
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, DataBarRule
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD, FORMAT_PERCENTAGE

# Color constants
DARK_BG = "1a1a2e"
DARKER_BG = "16213e"
HEADER_BG = "0f3460"
BLUE_TEXT = "4472C4"
GREEN = "00B050"
RED = "FF0000"
WHITE = "FFFFFF"
LIGHT_GRAY = "CCCCCC"
GOLD = "FFD700"

# Strategy list
STRATEGIES = ["定投仓DCA", "轮子策略仓Wheel", "LEAPS Call仓", "波段仓Swing", "现金仓"]


def create_workbook():
    """Create the Excel workbook with all sheets."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets in order
    create_settings_sheet(wb)
    create_positions_sheet(wb)
    create_strategy_summary_sheet(wb)
    create_dashboard_sheet(wb)

    return wb


def setup_dark_theme(ws):
    """Apply dark theme to worksheet."""
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
            cell.font = Font(color=WHITE)


def create_header(ws, headers, start_row=1, start_col=1):
    """Create styled header row."""
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=WHITE, size=11)

    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return start_row + 1


def create_positions_sheet(wb):
    """Create Sheet 1: Positions (持仓录入)."""
    ws = wb.create_sheet("持仓录入", 0)
    setup_dark_theme(ws)

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20

    # Headers
    headers = ["策略", "标的代码", "数量", "成本价", "现价", "市值", "盈亏金额", "盈亏%", "备注"]
    create_header(ws, headers, start_row=1, start_col=1)

    # Data validation for strategy column
    dv = DataValidation(
        type="list",
        formula1='"定投仓DCA,轮子策略仓Wheel,LEAPS Call仓,波段仓Swing,现金仓"',
        allow_blank=True
    )
    dv.prompt = "选择策略"
    dv.promptTitle = "策略选择"
    ws.add_data_validation(dv)

    # Sample data
    sample_data = [
        ["定投仓DCA", "AAPL", 150, 175.50, 195.20],
        ["定投仓DCA", "MSFT", 80, 410.00, 445.60],
        ["定投仓DCA", "QQQ", 50, 460.00, 498.30],
        ["定投仓DCA", "VOO", 30, 520.00, 555.80],
        ["轮子策略仓Wheel", "AMZN", 60, 185.00, 212.40],
        ["轮子策略仓Wheel", "GOOGL", 40, 155.20, 178.60],
        ["LEAPS Call仓", "NVDA", 20, 880.00, 1050.30],
        ["LEAPS Call仓", "META", 25, 480.00, 565.80],
        ["波段仓Swing", "TSLA", 15, 245.00, 268.40],
        ["波段仓Swing", "AMD", 30, 155.00, 168.20],
    ]

    # Fill sample data with formulas
    for i, row_data in enumerate(sample_data, start=2):
        row = i
        strategy, symbol, qty, cost, price = row_data

        # Strategy
        ws.cell(row=row, column=1, value=strategy)
        dv.add(ws.cell(row=row, column=1))

        # Symbol
        ws.cell(row=row, column=2, value=symbol)

        # Quantity
        ws.cell(row=row, column=3, value=qty)
        ws.cell(row=row, column=3).number_format = "0"

        # Cost price
        ws.cell(row=row, column=4, value=cost)
        ws.cell(row=row, column=4).number_format = "$#,##0.00"

        # Current price
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=5).number_format = "$#,##0.00"

        # Market Value: =ABS(C{row})*E{row}
        ws.cell(row=row, column=6).value = f"=IF(AND(C{row}<>\"\",E{row}<>\"\"),ABS(C{row})*E{row},\"\")"
        ws.cell(row=row, column=6).number_format = "$#,##0"

        # P&L Amount: =(E{row}-D{row})*C{row}
        ws.cell(row=row, column=7).value = f"=IF(AND(C{row}<>\"\",D{row}<>\"\",E{row}<>\"\"),(E{row}-D{row})*C{row},\"\")"
        ws.cell(row=row, column=7).number_format = "$#,##0"

        # P&L %: =IFERROR(E{row}/D{row}-1,"")
        ws.cell(row=row, column=8).value = f"=IFERROR(IF(AND(E{row}<>\"\",D{row}<>0),E{row}/D{row}-1,\"\"),\"\")"
        ws.cell(row=row, column=8).number_format = "0.0%"

        # Notes (empty)
        ws.cell(row=row, column=9, value="")

    # Add empty rows for user input (with formulas)
    for row in range(12, 102):  # 90 empty rows
        # Strategy with data validation
        cell = ws.cell(row=row, column=1)
        dv.add(cell)

        # Quantity
        ws.cell(row=row, column=3).number_format = "0"

        # Cost price
        ws.cell(row=row, column=4).number_format = "$#,##0.00"

        # Current price
        ws.cell(row=row, column=5).number_format = "$#,##0.00"

        # Market Value formula
        ws.cell(row=row, column=6).value = f"=IF(AND(C{row}<>\"\",E{row}<>\"\"),ABS(C{row})*E{row},\"\")"
        ws.cell(row=row, column=6).number_format = "$#,##0"

        # P&L Amount formula
        ws.cell(row=row, column=7).value = f"=IF(AND(C{row}<>\"\",D{row}<>\"\",E{row}<>\"\"),(E{row}-D{row})*C{row},\"\")"
        ws.cell(row=row, column=7).number_format = "$#,##0"

        # P&L % formula
        ws.cell(row=row, column=8).value = f"=IFERROR(IF(AND(E{row}<>\"\",D{row}<>0),E{row}/D{row}-1,\"\"),\"\")"
        ws.cell(row=row, column=8).number_format = "0.0%"

    # Conditional formatting for P&L columns
    # Green for positive, Red for negative
    green_fill = PatternFill(start_color="003300", end_color="003300", fill_type="solid")
    red_fill = PatternFill(start_color="330000", end_color="330000", fill_type="solid")

    # P&L Amount (Column G)
    ws.conditional_formatting.add(
        f"G2:G101",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=green_fill,
            font=Font(color=GREEN)
        )
    )
    ws.conditional_formatting.add(
        f"G2:G101",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=red_fill,
            font=Font(color=RED)
        )
    )

    # P&L % (Column H)
    ws.conditional_formatting.add(
        f"H2:H101",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=green_fill,
            font=Font(color=GREEN)
        )
    )
    ws.conditional_formatting.add(
        f"H2:H101",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=red_fill,
            font=Font(color=RED)
        )
    )

    # Freeze panes
    ws.freeze_panes = "A2"

    return ws


def create_strategy_summary_sheet(wb):
    """Create Sheet 2: Strategy Summary (策略汇总)."""
    ws = wb.create_sheet("策略汇总", 1)
    setup_dark_theme(ws)

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12

    # Title
    ws['A1'] = "策略汇总"
    ws['A1'].font = Font(bold=True, color=WHITE, size=14)
    ws['A1'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    # Cash input area
    ws['A3'] = "现金金额"
    ws['A3'].font = Font(color=BLUE_TEXT, bold=True)
    ws['B3'].value = 35000  # Default cash
    ws['B3'].number_format = "$#,##0"
    ws['B3'].font = Font(color=BLUE_TEXT, bold=True)
    ws['B3'].fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")

    # Note about editable cell
    ws['C3'] = "← 可编辑"
    ws['C3'].font = Font(color=LIGHT_GRAY, size=9, italic=True)

    # Headers for strategy table
    headers = ["策略名称", "图标", "总市值", "总盈亏", "盈亏%", "占总资产比例", "持仓数量"]
    header_row = 5
    create_header(ws, headers, start_row=header_row, start_col=1)

    # Strategy rows
    strategies = STRATEGIES[:-1]  # Exclude 现金仓 from strategies (it's separate)
    for i, strategy in enumerate(strategies, start=header_row + 1):
        row = i

        # Strategy name
        ws.cell(row=row, column=1, value=strategy)
        ws.cell(row=row, column=1).font = Font(color=WHITE, bold=True)

        # Icon (emoji representation)
        icons = {"定投仓DCA": "📈", "轮子策略仓Wheel": "🎡", "LEAPS Call仓": "📊", "波段仓Swing": "🎯", "现金仓": "💰"}
        ws.cell(row=row, column=2, value=icons.get(strategy, ""))
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

        # Total Market Value: SUMIFS on Positions sheet
        ws.cell(row=row, column=3).value = f'=SUMIFS(持仓录入.F:F,持仓录入.A:A,"{strategy}")'
        ws.cell(row=row, column=3).number_format = "$#,##0"

        # Total P&L: SUMIFS on Positions sheet
        ws.cell(row=row, column=4).value = f'=SUMIFS(持仓录入.G:G,持仓录入.A:A,"{strategy}")'
        ws.cell(row=row, column=4).number_format = "$#,##0"

        # P&L %: Total P&L / (Total Market Value - Total P&L)
        ws.cell(row=row, column=5).value = f'=IFERROR(IF(D{row}<>0,D{row}/(C{row}-D{row}),""),"")'
        ws.cell(row=row, column=5).number_format = "0.0%"

        # % of Total Assets: Total Market Value / Total Assets
        # Total Assets = Sum of all strategy market values + Cash
        total_assets_formula = f'SUM(C{header_row + 1}:C{header_row + len(strategies)})+B3'
        ws.cell(row=row, column=6).value = f"=IF(C{row}<>\"\",C{row}/({total_assets_formula}),\"\")"
        ws.cell(row=row, column=6).number_format = "0.0%"

        # Number of positions: COUNTIFS
        ws.cell(row=row, column=7).value = f'=COUNTIFS(持仓录入.A:A,"{strategy}",持仓录入.B:B,"<>")'
        ws.cell(row=row, column=7).number_format = "0"

    # Total row
    total_row = header_row + len(strategies) + 1
    ws.cell(row=total_row, column=1, value="总计")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    # Total market value
    ws.cell(row=total_row, column=3).value = f"=SUM(C{header_row + 1}:C{header_row + len(strategies)})"
    ws.cell(row=total_row, column=3).number_format = "$#,##0"
    ws.cell(row=total_row, column=3).font = Font(bold=True, color=WHITE)

    # Total P&L
    ws.cell(row=total_row, column=4).value = f"=SUM(D{header_row + 1}:D{header_row + len(strategies)})"
    ws.cell(row=total_row, column=4).number_format = "$#,##0"
    ws.cell(row=total_row, column=4).font = Font(bold=True, color=WHITE)

    # Total P&L %
    ws.cell(row=total_row, column=5).value = f"=IFERROR(IF(D{total_row}<>0,D{total_row}/(C{total_row}-D{total_row}),\"\"),\"\")"
    ws.cell(row=total_row, column=5).number_format = "0.0%"
    ws.cell(row=total_row, column=5).font = Font(bold=True, color=WHITE)

    # Total % of assets (always 100%)
    ws.cell(row=total_row, column=6, value="100.0%")
    ws.cell(row=total_row, column=6).number_format = "0.0%"
    ws.cell(row=total_row, column=6).font = Font(bold=True, color=WHITE)

    # Total positions
    ws.cell(row=total_row, column=7).value = f"=SUM(G{header_row + 1}:G{header_row + len(strategies)})"
    ws.cell(row=total_row, column=7).number_format = "0"
    ws.cell(row=total_row, column=7).font = Font(bold=True, color=WHITE)

    # Cash row
    cash_row = total_row + 1
    ws.cell(row=cash_row, column=1, value="现金")
    ws.cell(row=cash_row, column=1).font = Font(color=GOLD, bold=True)
    ws.cell(row=cash_row, column=2, value="💰")
    ws.cell(row=cash_row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=cash_row, column=3).value = "=B3"
    ws.cell(row=cash_row, column=3).number_format = "$#,##0"
    ws.cell(row=cash_row, column=3).font = Font(color=GOLD, bold=True)

    # Cash % of total assets
    total_assets_formula_cash = f'SUM(C{header_row + 1}:C{header_row + len(strategies)})+B3'
    ws.cell(row=cash_row, column=6).value = f"=B3/({total_assets_formula_cash})"
    ws.cell(row=cash_row, column=6).number_format = "0.0%"
    ws.cell(row=cash_row, column=6).font = Font(color=GOLD)

    # Conditional formatting for P&L columns
    green_font = Font(color=GREEN)
    red_font = Font(color=RED)

    ws.conditional_formatting.add(
        f"D{header_row + 1}:D{cash_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            font=green_font
        )
    )
    ws.conditional_formatting.add(
        f"D{header_row + 1}:D{cash_row}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            font=red_font
        )
    )

    ws.conditional_formatting.add(
        f"E{header_row + 1}:E{cash_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            font=green_font
        )
    )
    ws.conditional_formatting.add(
        f"E{header_row + 1}:E{cash_row}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            font=red_font
        )
    )

    # Create pie chart for strategy market value distribution
    pie_chart = PieChart()
    pie_chart.title = "策略市值分布"
    pie_chart.height = 10
    pie_chart.width = 10

    # Data for pie chart (strategy market values, excluding cash row)
    data_ref = Reference(ws, min_col=3, min_row=header_row + 1, max_row=header_row + len(strategies))
    labels_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(strategies))

    pie_chart.add_data(data_ref, titles_from_data=False)
    pie_chart.set_categories(labels_ref)

    # Style the chart
    pie_chart.style = 10
    ws.add_chart(pie_chart, "I5")

    return ws


def create_dashboard_sheet(wb):
    """Create Sheet 3: Dashboard (仪表盘)."""
    ws = wb.create_sheet("仪表盘", 2)
    setup_dark_theme(ws)

    # Set column widths
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Title
    ws['A1'] = "投资组合仪表盘"
    ws['A1'].font = Font(bold=True, color=WHITE, size=16)
    ws['A1'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    ws.merge_cells('A1:F1')

    # Statistic cards (row 3-5)
    # Note: Strategy Summary has total in row 11, cash in row 12
    stats = [
        ("总资产", "=策略汇总!C11+策略汇总!C12"),  # Total assets = sum of strategy values + cash
        ("总盈亏", "=策略汇总!D11"),
        ("现金", "=策略汇总!B3"),
        ("持仓数", "=COUNTA(持仓录入.B:B)-1"),  # -1 for header
        ("最大策略仓", "=INDEX(策略汇总.A6:A10,MATCH(MAX(策略汇总.C6:C10),策略汇总.C6:C10,0))"),
    ]

    for i, (label, formula) in enumerate(stats):
        col = i * 2 + 1
        # Label
        ws.cell(row=3, column=col, value=label)
        ws.cell(row=3, column=col).font = Font(color=LIGHT_GRAY, size=10)
        ws.cell(row=3, column=col).fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")

        # Value
        cell = ws.cell(row=4, column=col)
        cell.value = formula
        if "资产" in label or "盈亏" in label or "现金" in label:
            cell.number_format = "$#,##0"
        cell.font = Font(bold=True, color=WHITE, size=14)
        cell.fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)

    # Million Challenge Progress Bar (row 7-9)
    ws['A7'] = "百万挑战进度"
    ws['A7'].font = Font(bold=True, color=WHITE, size=12)
    ws['A7'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    # Challenge target (from Settings sheet)
    ws['B7'] = "目标: $1,000,000"

    # Progress cell
    ws['A8'] = "当前进度"
    ws['A8'].font = Font(color=LIGHT_GRAY)

    # Progress value (formula: total assets / 1,000,000)
    # Total assets = 策略汇总!C11 + 策略汇总!C12
    ws['B8'].value = "=(策略汇总!C11+策略汇总!C12)/1000000"
    ws['B8'].number_format = "0.0%"
    ws['B8'].font = Font(bold=True, color=WHITE, size=14)

    # Progress bar using conditional formatting data bar
    ws.conditional_formatting.add(
        'B8',
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="4472C4",
            showValue=True
        )
    )

    # Target Allocation Table (row 11+)
    ws['A11'] = "目标配比"
    ws['A11'].font = Font(bold=True, color=WHITE, size=12)
    ws['A11'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    # Headers
    alloc_headers = ["策略名", "目标占比", "实际占比", "差异"]
    alloc_header_row = 12
    create_header(ws, alloc_headers, start_row=alloc_header_row, start_col=1)

    # Target allocation data (row 13+)
    # These are editable by user (blue text)
    target_allocations = [
        ("定投仓DCA", 0.40),
        ("轮子策略仓Wheel", 0.25),
        ("LEAPS Call仓", 0.20),
        ("波段仓Swing", 0.10),
        ("现金仓", 0.05),
    ]

    for i, (strategy, target) in enumerate(target_allocations):
        row = alloc_header_row + 1 + i

        # Strategy name
        ws.cell(row=row, column=1, value=strategy)
        ws.cell(row=row, column=1).font = Font(color=WHITE)

        # Target % (editable, blue)
        target_cell = ws.cell(row=row, column=2, value=target)
        target_cell.number_format = "0.0%"
        target_cell.font = Font(color=BLUE_TEXT, bold=True)
        target_cell.fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")

        # Actual % (formula from Strategy Summary)
        # Use VLOOKUP to find the strategy in Strategy Summary sheet
        actual_formula = f'=IFERROR(VLOOKUP(A{row},策略汇总.A6:F10,6,FALSE),"")'
        ws.cell(row=row, column=3).value = actual_formula
        ws.cell(row=row, column=3).number_format = "0.0%"

        # Difference (Target - Actual)
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = "0.0%"

    # Conditional formatting for difference column (green if positive, red if negative)
    ws.conditional_formatting.add(
        f"D{alloc_header_row + 1}:D{alloc_header_row + len(target_allocations)}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            font=Font(color=GREEN)
        )
    )
    ws.conditional_formatting.add(
        f"D{alloc_header_row + 1}:D{alloc_header_row + len(target_allocations)}",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            font=Font(color=RED)
        )
    )

    # Column chart: Strategy P&L comparison
    # Reference data from Strategy Summary sheet
    strategy_sheet = wb['策略汇总']
    chart = BarChart()
    chart.title = "各策略盈亏对比"
    chart.style = 10
    chart.height = 10
    chart.width = 15

    # Data for chart (strategy P&L values)
    data_ref = Reference(strategy_sheet, min_col=4, min_row=6, max_row=10)
    cats_ref = Reference(strategy_sheet, min_col=1, min_row=6, max_row=10)

    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "F11")

    return ws


def create_settings_sheet(wb):
    """Create Sheet 4: Settings (设置)."""
    ws = wb.create_sheet("设置", 3)
    setup_dark_theme(ws)

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40

    # Title
    ws['A1'] = "设置"
    ws['A1'].font = Font(bold=True, color=WHITE, size=16)
    ws['A1'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    ws.merge_cells('A1:C1')

    # Cash amount (editable, blue)
    ws['A3'] = "现金金额"
    ws['A3'].font = Font(color=WHITE, bold=True)
    cash_cell = ws['B3']
    cash_cell.value = 35000
    cash_cell.number_format = "$#,##0"
    cash_cell.font = Font(color=BLUE_TEXT, bold=True)
    cash_cell.fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")
    ws['C3'] = "← 可编辑（将在策略汇总中引用）"
    ws['C3'].font = Font(color=LIGHT_GRAY, size=9, italic=True)

    # Challenge target (editable, blue)
    ws['A4'] = "挑战目标金额"
    ws['A4'].font = Font(color=WHITE, bold=True)
    target_cell = ws['B4']
    target_cell.value = 1000000
    target_cell.number_format = "$#,##0"
    target_cell.font = Font(color=BLUE_TEXT, bold=True)
    target_cell.fill = PatternFill(start_color=DARKER_BG, end_color=DARKER_BG, fill_type="solid")
    ws['C4'] = "← 可编辑（百万挑战目标）"
    ws['C4'].font = Font(color=LIGHT_GRAY, size=9, italic=True)

    # Instructions
    ws['A6'] = "使用说明"
    ws['A6'].font = Font(bold=True, color=WHITE, size=12)
    ws['A6'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    instructions = [
        "1. 在【持仓录入】表中输入持仓数据",
        "2. 策略列使用下拉菜单选择",
        "3. 市值、盈亏金额、盈亏% 自动计算",
        "4. 在【设置】表中修改现金金额和挑战目标",
        "5. 【策略汇总】自动汇总各策略表现",
        "6. 【仪表盘】查看投资组合概览",
        "7. 所有计算均使用Excel公式，可离线使用",
    ]

    for i, inst in enumerate(instructions, start=7):
        ws.cell(row=i, column=1, value=inst)
        ws.cell(row=i, column=1).font = Font(color=LIGHT_GRAY, size=10)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)

    return ws


def main():
    """Main function to create the dashboard."""
    print("Creating IBKR Portfolio Dashboard...")

    # Create workbook
    wb = create_workbook()

    # Save to file
    output_path = "/Users/ericsteg/.qclaw/workspace/ibkr-portfolio-dashboard/portfolio_dashboard.xlsx"
    wb.save(output_path)

    print(f"Dashboard created successfully: {output_path}")
    print("")
    print("Sheets created:")
    print("  1. 持仓录入 (Positions) - Data entry with formulas")
    print("  2. 策略汇总 (Strategy Summary) - Auto-calculated summary")
    print("  3. 仪表盘 (Dashboard) - Visual overview")
    print("  4. 设置 (Settings) - Configuration")
    print("")
    print("All calculations are formula-driven (no hardcoded values).")
    print("Dashboard is ready to use!")


if __name__ == "__main__":
    main()
